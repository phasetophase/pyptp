from __future__ import annotations

"""Internal helpers for reading Excel-based type sheets."""


from typing import TYPE_CHECKING

from openpyxl import load_workbook

from pyptp.ptp_log import logger

if TYPE_CHECKING:
    from collections.abc import Iterable


def _normalize_key(key: str) -> str:
    return key.casefold()


class TypeRow(dict[str, object]):
    """Row dict whose lookups ignore the case of the key.

    Workbook headers and the GNF/VNF property names that ``deserialize`` asks
    for differ in case throughout: the Cable sheet writes ``R_C`` and
    ``Shortname`` where the dataclasses read ``R_c`` and ``ShortName``.
    Resolving that here keeps ``column_renames`` for headers that are genuinely
    named something else.

    An exact match always wins; the case-insensitive lookup only applies when
    the exact key is absent.
    """

    def __init__(self, data: dict[str, object] | None = None) -> None:
        super().__init__(data or {})
        self._index = {_normalize_key(key): key for key in self}

    def _resolve(self, key: str) -> str:
        if dict.__contains__(self, key):
            return key
        return self._index.get(_normalize_key(key), key)

    def __contains__(self, key: object) -> bool:
        if not isinstance(key, str):
            return dict.__contains__(self, key)
        return dict.__contains__(self, self._resolve(key))

    def __getitem__(self, key: str) -> object:
        return dict.__getitem__(self, self._resolve(key))

    def __setitem__(self, key: str, value: object) -> None:
        dict.__setitem__(self, key, value)
        self._index[_normalize_key(key)] = key

    def __delitem__(self, key: str) -> None:
        actual = self._resolve(key)
        dict.__delitem__(self, actual)
        self._index.pop(_normalize_key(actual), None)

    def get(self, key: str, default: object = None) -> object:  # type: ignore[override]
        return dict.get(self, self._resolve(key), default)

    def pop(self, key: str, *default: object) -> object:
        actual = self._resolve(key)
        self._index.pop(_normalize_key(actual), None)
        return dict.pop(self, actual, *default)

    def copy(self) -> TypeRow:
        return TypeRow(self)


def read_sheet(
    path: str,
    sheet_name: str,
    *,
    skiprows: Iterable[int] | None = (),
) -> list[dict[str, object]]:
    """Read a sheet into a list of header->value row dicts; on failure returns an empty list."""
    skip = set(skiprows or ())
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        header: list[object] | None = None
        rows: list[dict[str, object]] = []
        for index, values in enumerate(ws.iter_rows(values_only=True)):
            if index in skip:
                continue
            if header is None:
                header = list(values)
                continue
            rows.append({str(col): val for col, val in zip(header, values, strict=False) if col is not None})
        return rows
    except Exception as exc:  # noqa: BLE001
        logger.debug("Failed reading sheet %s from %s: %s", sheet_name, path, exc)
        return []
    finally:
        if wb is not None:
            wb.close()


def normalize_rows(
    rows: list[dict[str, object]],
    rename: dict[str, str] | None = None,
) -> list[dict[str, object]]:
    """Drop all-empty rows and apply a key rename mapping.

    Rename keys match the source header ignoring case, the same way header
    lookup does, so a workbook is free to spell a column ``VoP`` or ``VOP``.
    """
    lookup = {_normalize_key(source): target for source, target in (rename or {}).items()}
    normalized: list[dict[str, object]] = []
    for row in rows:
        if all(value is None for value in row.values()):
            continue
        if lookup:
            normalized.append(TypeRow({lookup.get(_normalize_key(key), key): value for key, value in row.items()}))
        else:
            normalized.append(TypeRow(row))
    return normalized


def _drop_leading_unit_row(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Drop the first row when it is a unit row (empty ``Name``/``ShortName`` cell)."""
    if not rows:
        return rows
    first = rows[0]
    if "Name" in first:
        key = "Name"
    elif "ShortName" in first:
        key = "ShortName"
    else:
        return rows
    value = first[key]
    if value is None or not str(value).strip():
        return rows[1:]
    return rows


def read_type_sheet(
    path: str,
    sheet_name: str,
    *,
    rename: dict[str, str] | None = None,
) -> list[dict[str, object]]:
    """Read a type sheet into header->value row dicts, dropping a leading unit row.

    A unit row (``kV``, ``MVA``, ...) is detected by content (an empty
    ``Name``/``ShortName`` cell in the first row), not by position, so sheets
    without one keep their first data row.
    """
    rows = normalize_rows(read_sheet(path, sheet_name=sheet_name, skiprows=()), rename=rename)
    return _drop_leading_unit_row(rows)


def clean_row_dict(row: dict[str, object]) -> dict[str, object]:
    """Return a row dict with None values filtered out and keys coerced to str.

    Lookups on the result ignore case, so ``R_C`` in the workbook satisfies a
    ``data.get("R_c")`` in ``deserialize``.
    """
    return TypeRow({str(key): value for key, value in row.items() if value is not None})
